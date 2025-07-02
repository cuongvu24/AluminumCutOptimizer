import pandas as pd
import os
import json
import uuid
from datetime import datetime
from openpyxl.styles import PatternFill

def validate_input_excel(df):
    """Kiểm tra file nhập nhôm"""
    required = ["Mã Thanh", "Chiều Dài", "Số Lượng"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        return False, f"Thiếu cột: {', '.join(missing)}"

    try:
        df['Chiều Dài'] = pd.to_numeric(df['Chiều Dài'])
        df['Số Lượng'] = pd.to_numeric(df['Số Lượng'])
    except:
        return False, "Chiều Dài & Số Lượng phải là số"

    if (df['Chiều Dài'] <= 0).any():
        return False, "Chiều Dài phải > 0"
    if (df['Số Lượng'] <= 0).any():
        return False, "Số Lượng phải > 0"
    if df['Mã Thanh'].isnull().any():
        return False, "Mã Thanh không được để trống"

    return True, "Hợp lệ"

def create_accessory_summary(df, stream):
    required = ['Mã phụ kiện', 'Tên phụ phiện', 'Đơn vị tính', 'Số lượng']
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Thiếu cột: {', '.join(missing)}")
    grouped = df.groupby(['Mã phụ kiện', 'Tên phụ phiện', 'Đơn vị tính'])['Số lượng'].sum().reset_index()
    grouped = grouped.rename(columns={'Số lượng': 'Tổng Số Lượng'})
    grouped.to_excel(stream, sheet_name="Tổng Hợp Phụ Kiện", index=False)
    return grouped

def create_output_excel(stream, result_df, patterns_df, summary_df, stock_length_options, cutting_gap, unassigned_df=None):
    with pd.ExcelWriter(stream, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name="Tổng Hợp", index=False)
        patterns_df.to_excel(writer, sheet_name="Mẫu Cắt", index=False)
        result_df.to_excel(writer, sheet_name="Chi Tiết Mảnh", index=False)
        if unassigned_df is not None and not unassigned_df.empty:
            unassigned_df.to_excel(writer, sheet_name="Sót Lại", index=False)

        # Sheet Tham Số
        params = pd.DataFrame({
            'Tham Số': ['Kích Thước Thanh', 'Khoảng Cách Cắt'],
            'Giá Trị': [', '.join(map(str, stock_length_options)), cutting_gap]
        })
        params.to_excel(writer, sheet_name="Tham Số", index=False)

        # Sheet mô phỏng
        try:
            ws = writer.book.create_sheet("Mô Phỏng Cắt")
            piece_colors = ["FF9999", "99FF99", "9999FF", "FFFF99", "FF99FF", "99FFFF"]

            if not patterns_df.empty:
                patterns_df = patterns_df.sort_values('Mã Thanh')
                max_pieces = patterns_df['Mẫu Cắt'].apply(lambda x: len(x.split('+'))).max()
                headers = [col for col in patterns_df.columns if col != 'Mẫu Cắt'] + [f'Đoạn {i+1}' for i in range(max_pieces)]
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)

                for idx, row in enumerate(patterns_df.itertuples(index=False), 2):
                    for col_idx, col in enumerate([c for c in patterns_df.columns if c != 'Mẫu Cắt'], 1):
                        ws.cell(row=idx, column=col_idx, value=getattr(row, col))

                    pieces = getattr(row, 'Mẫu Cắt').split('+')
                    for piece_idx, piece in enumerate(pieces):
                        cidx = len(patterns_df.columns) + piece_idx
                        cell = ws.cell(row=idx, column=cidx, value=float(piece))
                        cell.number_format = "0.0"
                        fill = PatternFill(start_color=piece_colors[piece_idx % len(piece_colors)],
                                           end_color=piece_colors[piece_idx % len(piece_colors)],
                                           fill_type="solid")
                        cell.fill = fill

            else:
                ws.cell(row=1, column=1, value="Không có dữ liệu mô phỏng")
        except Exception as e:
            ws = writer.book.create_sheet("Mô Phỏng Cắt")
            ws.cell(row=1, column=1, value=f"Lỗi: {str(e)}")

def save_optimization_history(result_df, patterns_df, summary_df, stock_length_options, cutting_gap, method, name=None):
    f = "history.json"
    history = []
    if os.path.exists(f):
        try:
            with open(f, "r", encoding="utf-8") as r:
                history = json.load(r)
        except:
            history = []

    entry = {
        "id": str(uuid.uuid4()),
        "name": name or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": method,
        "stock_length_options": stock_length_options,
        "cutting_gap": cutting_gap,
        "profiles": summary_df['Mã Thanh'].tolist(),
        "result_df": result_df.to_dict(),
        "patterns_df": patterns_df.to_dict(),
        "summary_df": summary_df.to_dict()
    }

    history.append(entry)
    with open(f, "w", encoding="utf-8") as w:
        json.dump(history, w, ensure_ascii=False, indent=2)

def load_optimization_history():
    if os.path.exists("history.json"):
        try:
            with open("history.json", "r", encoding="utf-8") as r:
                return json.load(r)
        except:
            return []
    return []

def delete_optimization_history_entry(id):
    if os.path.exists("history.json"):
        try:
            with open("history.json", "r", encoding="utf-8") as r:
                h = json.load(r)
            h = [i for i in h if i['id'] != id]
            with open("history.json", "w", encoding="utf-8") as w:
                json.dump(h, w, ensure_ascii=False, indent=2)
        except:
            pass
